# scanner.py
# DevBraze QR Scanner — polished UI + fixed camera frame clipping (no overlay bleed),
# rear camera preference, manual switch, optional sounds, no gate field.

import os
from datetime import datetime
from zoneinfo import ZoneInfo

from flask import Flask, request, jsonify, render_template_string, send_from_directory
from openpyxl import load_workbook
from filelock import FileLock

# ========= CONFIG =========
APP_DIR     = os.path.dirname(__file__)
ASSETS_DIR  = os.path.join(APP_DIR, "assets")
EXCEL_PATH  = os.path.join(ASSETS_DIR, "Teams.xlsx")  # must exist; sheet "Teams"
SHEET_NAME  = "Teams"
TIMEZONE    = "Asia/Kolkata"
REQUIRED_COLS = ["Token", "Entry Confirmed", "Check-in Time", "Check-in Gate"]  # Gate kept for compatibility
# =========================

app = Flask(__name__)

SCANNER_HTML = """
<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1"/>
<link rel="icon" href="/assets/devbraze_logo.png">
<title>DevBraze — Code Crusade Scanner</title>
<style>
  :root {
    --brand:#0F766E; --accent:#0EA5E9; --teal:#14b8a6;
    --bg:#f6f7f9; --card:#ffffff; --text:#111827; --muted:#6b7280;
  }
  * { box-sizing:border-box; }
  html,body { height:100%; }
  body { font-family: system-ui, -apple-system, Segoe UI, Roboto, Arial, sans-serif; margin:0; background:var(--bg); color:var(--text); }

  /* Header */
  header {
    background: linear-gradient(90deg, var(--brand) 0%, var(--accent) 100%);
    color:white; padding:14px 20px; font-size:18px; font-weight:700;
    display:flex; align-items:center; justify-content:space-between;
    border-bottom:3px solid var(--teal);
  }
  header .title { display:flex; gap:10px; align-items:center; }
  header img { height:36px; border-radius:8px; background:rgba(255,255,255,0.10); padding:4px; }

  /* Layout */
  main { max-width:1100px; margin:24px auto 80px; padding:0 16px; }
  .card {
    background: rgba(255,255,255,0.85);
    backdrop-filter: blur(8px);
    border: 1px solid rgba(255,255,255,0.35);
    border-radius: 18px;
    box-shadow: 0 14px 34px rgba(0,0,0,.10);
    padding: 18px;
    animation: fadeIn .4s ease;
  }
  @keyframes fadeIn { from{opacity:0; transform:translateY(8px);} to{opacity:1; transform:translateY(0);} }

  .row { display:flex; gap:20px; flex-wrap:wrap; align-items:stretch; }
  .left { min-width:320px; }
  .right { flex:1; min-width:300px; display:flex; flex-direction:column; }

  /* Camera frame with gradient border that CLIPS inner content (fixes overlay bleed) */
  #reader {
    position: relative;
    width: 420px;                 /* camera box width (auto on mobile via max-width) */
    max-width: 100%;
    border-radius: 16px;
    overflow: hidden;             /* clip inner video/overlay corners */
    background: #fff;             /* inner background */
    box-shadow: 0 8px 24px rgba(14,165,233,0.18);
  }
  /* Gradient border without using padding (prevents overlay from escaping) */
  #reader::before {
    content: "";
    position: absolute;
    inset: 0;
    border-radius: inherit;
    padding: 3px;                 /* border thickness */
    background: linear-gradient(90deg, var(--brand), var(--accent));
    -webkit-mask:
      linear-gradient(#000 0 0) content-box,
      linear-gradient(#000 0 0);
    -webkit-mask-composite: xor;
            mask-composite: exclude;
    pointer-events: none;
  }
  /* Ensure nested canvases also clip */
  #reader > * { position: relative; border-radius: inherit; overflow: hidden; }

  /* Controls */
  .toolbar { display:flex; gap:8px; align-items:center; flex-wrap:wrap; margin-bottom:12px; }
  label { font-size:14px; color:#374151; }
  select, input { padding:10px 12px; border-radius:12px; border:1px solid #e5e7eb; font-size:14px; background:white; }
  button {
    padding:10px 14px; border-radius:12px; border:0; cursor:pointer; font-weight:700; color:white;
    background: var(--accent);
  }
  button:hover { filter:brightness(1.05); }
  #btnSwitch {
    border-radius: 999px;
    background: linear-gradient(90deg, #0EA5E9, #38bdf8);
    box-shadow: 0 3px 10px rgba(14,165,233,0.35);
  }
  .secondary { background:#374151; }

  .grid { display:grid; grid-template-columns: 1fr; gap:10px; margin-top:12px; }

  /* Status + log */
  .status { min-height:46px; display:flex; align-items:center; padding:12px; border-radius:12px; font-weight:700; }
  .ok   { background:#ecfdf5; color:#065f46; border:1px solid #a7f3d0; }
  .warn { background:#fef3c7; color:#92400e; border:1px solid #fde68a; }
  .err  { background:#fee2e2; color:#991b1b; border:1px solid #fecaca; }
  #stats {
    margin-top:6px; display:inline-block; align-self:flex-end;
    background:#ecfdf5; border:1px solid #a7f3d0; color:#065f46;
    padding:4px 10px; border-radius:12px; font-weight:700; font-size:14px;
  }
  .log { margin-top:8px; flex:1; background:#0f172a; color:#e5e7eb; padding:12px; border-radius:12px; font-family: ui-monospace, SFMono-Regular, Menlo, Consolas, monospace; font-size:12px; overflow:auto; }

  /* Flash feedback */
  .flash-success { animation: flashGreen 260ms ease; }
  .flash-warn    { animation: flashYellow 260ms ease; }
  .flash-error   { animation: flashRed 260ms ease; }
  @keyframes flashGreen { 0%{box-shadow:0 0 24px 8px #22c55e;} 100%{box-shadow:none;} }
  @keyframes flashYellow{ 0%{box-shadow:0 0 24px 8px #facc15;} 100%{box-shadow:none;} }
  @keyframes flashRed   { 0%{box-shadow:0 0 24px 8px #ef4444;} 100%{box-shadow:none;} }

  /* Loading overlay */
  #loading {
    position:fixed; inset:0; display:flex; align-items:center; justify-content:center;
    background:rgba(0,0,0,.55); color:white; z-index:999; backdrop-filter: blur(2px);
    font-weight:700; letter-spacing:.3px;
  }
  .pill { background:rgba(255,255,255,.12); padding:10px 14px; border-radius:999px; border:1px solid rgba(255,255,255,.2); }

  /* Footer badge */
  footer {
    position:fixed; bottom:12px; right:16px;
    background:var(--brand); color:white; padding:6px 12px; border-radius:18px;
    font-size:13px; opacity:0.95; box-shadow:0 4px 14px rgba(0,0,0,.15);
  }

  @media (max-width: 720px) {
    main { margin-top:16px; }
    .row { flex-direction:column; }
    #reader { width:100%!important; }
    .card { padding:14px; }
    header { padding:12px 16px; font-size:16px; }
  }
</style>
</head>
<body>
<div id="loading"><div class="pill">🎥 Initializing camera…</div></div>

<header>
  <div class="title">Code Crusade Hackathon — Check-In</div>
  <img src="/assets/devbraze_logo.png" alt="DevBraze"/>
</header>

<main>
  <div class="card">
    <div class="row">
      <div class="left">
        <div class="toolbar">
          <label for="cameraSelect" style="margin-right:6px;">Camera</label>
          <select id="cameraSelect"></select>
          <button id="btnSwitch">Switch</button>
        </div>

        <div id="reader"></div>

        <div class="grid">
          <label>Manual token (fallback)</label>
          <input id="manual" placeholder="Paste/enter token here"/>
          <button id="btnManual">Check-in Manually</button>
        </div>
      </div>

      <div class="right">
        <div id="status" class="status warn">Ready. Scan a QR or enter a token.</div>
        <div id="stats">✅ 0 checked-in (session)</div>
        <div class="log" id="log"></div>
        <div style="margin-top:8px;">
          <button id="clearLog" class="secondary" type="button">Clear Log</button>
        </div>
      </div>
    </div>
  </div>
</main>

<footer>© DevBraze Club — REVA University</footer>



<script src="https://unpkg.com/html5-qrcode"></script>
<script>
const statusBox   = document.getElementById('status');
const logBox      = document.getElementById('log');
const statsBox    = document.getElementById('stats');
const manualEl    = document.getElementById('manual');
const btnManual   = document.getElementById('btnManual');
const cameraSel   = document.getElementById('cameraSelect');
const btnSwitch   = document.getElementById('btnSwitch');
const clearLogBtn = document.getElementById('clearLog');
const loading     = document.getElementById('loading');

const sndOk       = document.getElementById('sound-ok');
const sndRepeat   = document.getElementById('sound-repeat');
const sndError    = document.getElementById('sound-error');

let inFlight = false;
let reader   = null;
let cameras  = [];
let currentId = null;

let countChecked = 0;   // session count
let totalTeams   = null;

function safePlay(el){ try { el && el.play().catch(()=>{}); } catch(_){} }

function flash(kind){
  const cls = kind === 'ok' ? 'flash-success' : kind === 'warn' ? 'flash-warn' : 'flash-error';
  document.body.classList.add(cls);
  setTimeout(()=>document.body.classList.remove(cls), 300);
}

function setStatus(kind, msg){
  statusBox.className = 'status ' + kind;
  statusBox.textContent = msg;
  logBox.textContent = `[${new Date().toLocaleTimeString()}] ${msg}\\n` + logBox.textContent;
}

function updateStats(){
  statsBox.textContent = totalTeams != null
    ? `✅ ${countChecked} / ${totalTeams} Teams Checked-In`
    : `✅ ${countChecked} checked-in (session)`;
}

async function checkin(token){
  if (inFlight) return;
  inFlight = true;
  try{
    const r = await fetch('/api/checkin', {
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body: JSON.stringify({ token })
    });
    let text = await r.text();
    let j = {}; try { j = JSON.parse(text); } catch {}

    if (!r.ok) {
      const msg = j.message || text || 'Server error';
      setStatus('err', `❌ ${msg.trim()}`); flash('err'); safePlay(sndError);
    } else if (j.status === 'ok') {
      setStatus('ok', `✅ Checked in at ${j.time}`); flash('ok'); safePlay(sndOk);
      countChecked++; updateStats();
    } else if (j.status === 'repeat') {
      setStatus('warn', `⚠️ Already checked in at ${j.time}`); flash('warn'); safePlay(sndRepeat);
    } else {
      setStatus('err', `❌ ${j.message || 'Invalid response'}`); flash('err'); safePlay(sndError);
    }
  } catch(e){
    setStatus('err', 'Network or server error.'); flash('err'); safePlay(sndError);
  } finally {
    inFlight = false;
  }
}

btnManual.onclick = () => {
  const t = manualEl.value.trim();
  if (!t) return setStatus('warn','Enter a token first.');
  checkin(t);
};
clearLogBtn.onclick = () => { logBox.textContent = ''; };

function onScanSuccess(decodedText){
  const token = decodedText.trim();
  if (!token) return;
  checkin(token);
}
function onScanFailure(_){ /* ignore */ }

async function stopReader(){
  if (reader) {
    try { await reader.stop(); } catch(_) {}
    try { await reader.clear(); } catch(_) {}
  }
}

/* Compute a qrbox that fits INSIDE the rounded frame so the guides don't hit the border */
function computeQrbox() {
  const el = document.getElementById('reader');
  const w = el.getBoundingClientRect().width;
  const margin = 32;                                // safe margin from edges
  return Math.max(180, Math.min(340, Math.floor(w - margin)));
}

async function startReader(deviceId){
  await stopReader();
  reader = new Html5Qrcode("reader");
  currentId = deviceId;
  try{
    await reader.start(
      deviceId,
      { fps: 10, qrbox: computeQrbox() },           // dynamic qrbox
      onScanSuccess,
      onScanFailure
    );
    setStatus('ok', 'Camera running');
  } catch(e){
    setStatus('err', 'Failed to start camera. Try another option.');
  } finally {
    loading.style.display = 'none';
  }
}

function pickRearCamera(list){
  const pref = ['back','rear','environment','trás','arrière','后置','背面'];
  const lc = s => (s||'').toLowerCase();
  for (const w of pref){
    const hit = list.find(c => lc(c.label).includes(w));
    if (hit) return hit.id;
  }
  if (list.length >= 2) return list[list.length - 1].id; // often rear
  return list[0]?.id || null;
}

function populateCameraDropdown(list, preferredId){
  cameraSel.innerHTML = '';
  list.forEach((c, i) => {
    const opt = document.createElement('option');
    opt.value = c.id;
    opt.textContent = c.label || `Camera ${i+1}`;
    if (c.id === preferredId) opt.selected = true;
    cameraSel.appendChild(opt);
  });
}

btnSwitch.onclick = async () => {
  const chosen = cameraSel.value;
  if (!chosen) return;
  loading.style.display = 'flex';
  await startReader(chosen);
};

(async function init(){
  // Nudge permission on some browsers so labels appear
  try{
    const tmp = new Html5Qrcode("reader");
    const cams = await Html5Qrcode.getCameras();
    const first = cams[0]?.id;
    if (first) { try { await tmp.start(first, {fps:1}, ()=>{}, ()=>{}); } catch(_) {} }
    try { await tmp.stop(); } catch(_) {}
    try { await tmp.clear(); } catch(_) {}
  } catch(_) {}

  try{
    cameras = await Html5Qrcode.getCameras();
    if (!cameras || !cameras.length){
      setStatus('err','No camera available. Use manual token entry.');
      loading.style.display = 'none';
      return;
    }
    const preferred = pickRearCamera(cameras);
    populateCameraDropdown(cameras, preferred);
    await startReader(preferred);
  } catch(e){
    setStatus('err','Camera enumeration failed. Use manual token entry.');
    loading.style.display = 'none';
  }
  updateStats();

  // Recompute qrbox on resize for better fit
  window.addEventListener('resize', () => {
    // best-effort: if running, restart with new qrbox (lightweight on mobile)
    if (currentId && reader) {
      startReader(currentId);
    }
  });
})();
</script>
</body>
</html>
"""

# ---------- helpers ----------
def ensure_excel_columns(ws):
    header = {(ws.cell(row=1, column=col).value or "").strip(): col
              for col in range(1, ws.max_column + 1)}
    for name in REQUIRED_COLS:
        if name not in header:
            ws.cell(row=1, column=ws.max_column + 1, value=name)
    header = {(ws.cell(row=1, column=col).value or "").strip(): col
              for col in range(1, ws.max_column + 1)}
    return header

def now_local_string():
    try:
        tz = ZoneInfo(TIMEZONE)
        return datetime.now(tz).strftime("%Y-%m-%d %H:%M:%S %Z")
    except Exception:
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def safe_save_workbook(wb, target_path: str):
    try:
        wb.save(target_path)
        return {"saved_to": target_path, "autosave": False, "error": None}
    except PermissionError as e:
        base, ext = os.path.splitext(target_path)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        alt = f"{base}.autosave_{stamp}{ext}"
        wb.save(alt)
        return {"saved_to": alt, "autosave": True, "error": str(e)}

# ---------- routes ----------
@app.get("/scanner")
def scanner():
    return render_template_string(SCANNER_HTML)

@app.get("/")
def root():
    return "<meta http-equiv='refresh' content='0; url=/scanner' />"

# Serve /assets/* (logo, sounds, Excel if needed)
@app.get("/assets/<path:filename>")
def serve_assets(filename):
    return send_from_directory(ASSETS_DIR, filename)

@app.post("/api/checkin")
def api_checkin():
    try:
        data = request.get_json(force=True, silent=True) or {}
        token = (data.get("token") or "").strip()
        if not token:
            return jsonify(status="error", message="No token provided"), 400

        if not os.path.exists(EXCEL_PATH):
            return jsonify(status="error", message=f"Excel not found: {EXCEL_PATH}"), 500

        lock = FileLock(EXCEL_PATH + ".lock")
        with lock:
            wb = load_workbook(EXCEL_PATH)
            try:
                if SHEET_NAME not in wb.sheetnames:
                    return jsonify(status="error", message=f"Sheet '{SHEET_NAME}' not found"), 500
                ws = wb[SHEET_NAME]

                header = ensure_excel_columns(ws)
                col_token   = header["Token"]
                col_status  = header["Entry Confirmed"]
                col_time    = header["Check-in Time"]

                # find row by token
                hit_row = None
                for r in range(2, ws.max_row + 1):
                    cell_val = ws.cell(row=r, column=col_token).value
                    if (str(cell_val or "").strip()) == token:
                        hit_row = r
                        break
                if not hit_row:
                  return jsonify(status="error", message="Invalid token"), 404

                team_id = (str(ws.cell(row=hit_row, column=1).value or "").strip())
                current = (str(ws.cell(row=hit_row, column=col_status).value or "").strip())
                time_val = ws.cell(row=hit_row, column=col_time).value

                if current and current.lower() in ("yes", "confirmed", "present"):
                    return jsonify(status="repeat", team_id=team_id, time=str(time_val) if time_val else ""), 200

                ws.cell(row=hit_row, column=col_status, value="Yes")
                ts = now_local_string()
                ws.cell(row=hit_row, column=col_time, value=ts)

                result = safe_save_workbook(wb, EXCEL_PATH)
                payload = {"status": "ok", "team_id": team_id, "time": ts}
                if result["autosave"]:
                    payload["note"] = f"Original file locked. Wrote to '{os.path.basename(result['saved_to'])}'. Merge later."
                return jsonify(**payload), 200
            finally:
                wb.close()

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify(status="error", message=f"Server exception: {type(e).__name__}: {e}"), 500

# ---------- main ----------
if __name__ == "__main__":
    port = int(os.getenv("PORT", "5000"))
    app.run(host="0.0.0.0", port=port, debug=False)
