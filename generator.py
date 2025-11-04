"""
Team Ticket Generator (QR = TOKEN ONLY) — Styled, space-filled layout

What it does
------------
- Reads: Teams.xlsx (sheet "Teams") with at least the column: "Team ID"
- Adds/uses columns: Token, Entry Confirmed, Check-in Time
- Creates:
    tickets/png/<token>.png   (QR image containing just the token text)
    tickets/pdf/<Team ID>.pdf (printable, nicely designed ticket)

Install:
    pip install qrcode[pil] reportlab pillow openpyxl

Run:
    python generator.py
"""

import os, re, uuid
from pathlib import Path
from openpyxl import load_workbook
import qrcode
from reportlab.lib.pagesizes import A6, landscape
from reportlab.pdfgen import canvas
from reportlab.lib.colors import HexColor, Color

# ====== CONFIG ======
EXCEL_PATH    = r"C:\Users\exion\Downloads\Teams.xlsx"   # must exist; sheet must contain "Team ID"
SHEET_NAME    = "Teams"
OUT_DIR       = "tickets"

# Branding / copy
EVENT_NAME    = "Code Crusade Hackathon"     # top colored header
CLUB_NAME     = "Devbraze Community"         # shown above QR inside white area
EVENT_DATE    = "Nov 23–24, 2025"
EVENT_VENUE   = "REVA University"
PASS_TYPE     = "Team Pass"

# Colors
BRAND_PRIMARY = "#0F766E"  # teal band
BRAND_ACCENT  = "#0EA5E9"  # highlight text
BORDER_GRAY   = "#e5e7eb"  # light borders
TEXT_MUTED    = "#4b5563"  # gray-600
CHIP_BORDER   = "#d1d5db"  # gray-300

# Optional logo (PNG/JPG). Use a raw string or forward slashes on Windows.
LOGO_PATH     = r"C:\Users\exion\OneDrive\Desktop\LOGOS\DevBraze Logo Blue Border.png"  # or None
# =====================

ADDED_COLS = ["Token", "Entry Confirmed", "Check-in Time"]


# ---------- helpers ----------
def safe_name(s: str) -> str:
    s = str(s)
    s = re.sub(r'[\\/:*?"<>|]+', "_", s)
    s = s.strip()
    return s or "unnamed"

def ensure_columns(ws):
    header = {(ws.cell(row=1, column=c).value or "").strip(): c
              for c in range(1, ws.max_column + 1)}
    if "Team ID" not in header:
        raise SystemExit("Missing required column 'Team ID' in sheet 'Teams'.")

    for col in ADDED_COLS:
        if col not in header:
            ws.cell(row=1, column=ws.max_column + 1, value=col)

    # rebuild header map after additions
    header = {(ws.cell(row=1, column=c).value or "").strip(): c
              for c in range(1, ws.max_column + 1)}
    return header

def make_qr_token(token: str, outpath: str):
    os.makedirs(os.path.dirname(outpath), exist_ok=True)
    qr = qrcode.QRCode(box_size=8, border=2)
    qr.add_data(token)  # ONLY THE TOKEN, no URL
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    img.save(outpath)
    return outpath

def _rounded_rect(c, x, y, w, h, r=10, fill=0, stroke=1):
    c.roundRect(x, y, w, h, r, fill=fill, stroke=stroke)

def _chip(c, x, y, text, pad_x=8):
    c.setFont("Helvetica", 9)
    tw = c.stringWidth(text, "Helvetica", 9)
    w = tw + pad_x * 2
    h = 18
    c.setFillColor(Color(0,0,0,alpha=0))
    c.setStrokeColor(HexColor(CHIP_BORDER))
    c.setLineWidth(1)
    _rounded_rect(c, x, y, w, h, r=9, fill=0, stroke=1)
    c.setFillColor(HexColor(TEXT_MUTED))
    c.drawString(x + pad_x, y + 5, text)
    return w, h

# ---------- ticket layout ----------
def draw_pdf(team_id: str, token: str, qr_png: str, pdf_path: str):
    os.makedirs(os.path.dirname(pdf_path), exist_ok=True)
    c = canvas.Canvas(pdf_path, pagesize=landscape(A6))
    w, h = landscape(A6)

    primary = HexColor(BRAND_PRIMARY)
    accent  = HexColor(BRAND_ACCENT)
    border  = HexColor(BORDER_GRAY)
    muted   = HexColor(TEXT_MUTED)

    # ---------- Header (event name only) ----------
    band_h = 36
    c.setFillColor(primary)
    c.rect(0, h - band_h, w, band_h, fill=1, stroke=0)
    c.setFillColor("white")
    c.setFont("Helvetica-Bold", 16)
    c.drawCentredString(w/2, h - band_h/2 - 6, EVENT_NAME)

    # ---------- Card ----------
    card_m   = 12
    card_x   = card_m
    card_y   = 12
    card_w   = w - 2*card_m
    card_h   = h - band_h - card_m - card_y
    c.setFillColor("white")
    c.setStrokeColor(border)
    c.setLineWidth(1)
    c.roundRect(card_x, card_y, card_w, card_h, 14, fill=1, stroke=1)

    # inner padding
    pad      = 16
    inner_x  = card_x + pad
    inner_y  = card_y + pad
    inner_w  = card_w - 2*pad
    inner_h  = card_h - 2*pad

    # ---------- Club row (perfectly aligned) ----------
    row_top  = card_y + card_h - pad  # top inner edge
    # baseline for club text/logo
    club_baseline = row_top - 10

    c.setFillColor("black")
    c.setFont("Helvetica-Bold", 12)
    c.drawString(inner_x, club_baseline, CLUB_NAME)

    # logo aligned to right on same baseline
    logo_h = 18
    logo_w = 18
    logo_x = card_x + card_w - pad - logo_w
    logo_y = club_baseline - (logo_h - 10)  # match text baseline
    if LOGO_PATH and os.path.exists(LOGO_PATH):
        try:
            c.drawImage(LOGO_PATH, logo_x, logo_y, width=logo_w, height=logo_h,
                        preserveAspectRatio=True, mask="auto")
        except Exception:
            c.setStrokeColor(border); c.rect(logo_x, logo_y, logo_w, logo_h, fill=0, stroke=1)
    else:
        c.setStrokeColor(border); c.rect(logo_x, logo_y, logo_w, logo_h, fill=0, stroke=1)

    # hairline divider under club row
    c.setStrokeColor(border); c.setLineWidth(1)
    divider_y = club_baseline - 12
    c.line(inner_x, divider_y, card_x + card_w - pad, divider_y)

    # ---------- Main content grid ----------
    # QR column
    qr_size = 140
    qr_x = inner_x
    # center QR vertically in remaining area
    body_top = divider_y - 8
    body_bottom = inner_y
    body_h = body_top - body_bottom
    qr_y = body_bottom + (body_h - qr_size) / 2
    c.drawImage(qr_png, qr_x, qr_y, width=qr_size, height=qr_size, preserveAspectRatio=True)

    # Vertical divider centered to QR block
    div_x = qr_x + qr_size + 18
    c.setStrokeColor(border); c.setLineWidth(1)
    c.line(div_x, qr_y, div_x, qr_y + qr_size)

    # Text column aligned to QR middle
    txt_x = div_x + 16
    txt_mid_y = qr_y + qr_size/2

    # TEAM line
    c.setFillColor("black"); c.setFont("Helvetica-Bold", 16)
    c.drawString(txt_x, txt_mid_y + 30, f"TEAM: {team_id}")

    # TOKEN line
    c.setFont("Helvetica", 10); c.setFillColor(accent)
    c.drawString(txt_x, txt_mid_y + 10, f"TOKEN: {token[:8]}…")

    # Instruction
    c.setFillColor(muted); c.setFont("Helvetica", 9)
    c.drawString(txt_x, txt_mid_y - 8, "Present this ticket at entry")

    # Chips row (kept within card width)
    chip_y = qr_y + 8
    c.setFont("Helvetica", 9); c.setFillColor(muted)
    # simple chip function inline for precise spacing
    def chip(x, text):
        tw = c.stringWidth(text, "Helvetica", 9)
        w_chip = tw + 16
        c.setStrokeColor(HexColor(CHIP_BORDER)); c.setLineWidth(1)
        c.roundRect(x, chip_y, w_chip, 18, 9, fill=0, stroke=1)
        c.drawString(x + 8, chip_y + 5, text)
        return w_chip + 8

    cx = txt_x
    cx += chip(cx, EVENT_DATE)
    cx += chip(cx, EVENT_VENUE)
    # omit "Powered by ..." per your request (no footer)

    c.showPage()
    c.save()



# ---------- main ----------
def main():
    excel = Path(EXCEL_PATH)
    if not excel.exists():
        raise SystemExit(f"Excel not found: {EXCEL_PATH}")

    wb = load_workbook(excel)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Sheet '{SHEET_NAME}' not found.")
    ws = wb[SHEET_NAME]

    header = ensure_columns(ws)
    teamid_col    = header["Team ID"]
    token_col     = header["Token"]
    confirmed_col = header["Entry Confirmed"]
    time_col      = header["Check-in Time"]

    os.makedirs(f"{OUT_DIR}/png", exist_ok=True)
    os.makedirs(f"{OUT_DIR}/pdf", exist_ok=True)

    made = 0
    for r in range(2, ws.max_row + 1):
        team_id_raw = ws.cell(row=r, column=teamid_col).value
        if team_id_raw is None or str(team_id_raw).strip() == "":
            continue
        team_id = str(team_id_raw).strip()

        # generate or reuse token
        token_cell = ws.cell(row=r, column=token_col)
        token = str(token_cell.value or "").strip()
        if not token:
            token = str(uuid.uuid4())
            token_cell.value = token

        # initialize status columns if empty
        if ws.cell(row=r, column=confirmed_col).value is None:
            ws.cell(row=r, column=confirmed_col).value = ""
        if ws.cell(row=r, column=time_col).value is None:
            ws.cell(row=r, column=time_col).value = ""

        # QR contains only the token
        qr_png = f"{OUT_DIR}/png/{token}.png"
        make_qr_token(token, qr_png)

        # PDF per team (safe filename)
        pdf_path = f"{OUT_DIR}/pdf/{safe_name(team_id)}.pdf"
        draw_pdf(team_id, token, qr_png, pdf_path)
        made += 1
        print(f"[OK] {team_id} → {pdf_path}")

    wb.save(excel)
    wb.close()
    print(f"\nGenerated {made} team tickets. PDFs in '{OUT_DIR}/pdf/'.")
    print("Each QR encodes ONLY the token. Tokens saved under 'Token' column.")


if __name__ == "__main__":
    main()
